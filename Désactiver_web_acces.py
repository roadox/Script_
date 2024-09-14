import openpyxl
from netmiko import ConnectHandler

# Informations de connexion au commutateur Cisco (utilisateur et mot de passe communs)
username = "Username"
password = "MotDePasse"

# Ouvrir le fichier Excel
workbook = openpyxl.load_workbook("inventaire_switch_0day_Copie.xlsx")
sheet = workbook.active

# Parcourir la colonne A du fichier Excel et remplir la colonne D
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Commence à la deuxième ligne pour ignorer l'en-tête
    ip_address = row[0]

    print(f"Tentative de connexion à l'adresse IP : {ip_address}")

    # Informations de connexion au commutateur Cisco
    device = {
        'device_type': 'cisco_ios',
        'ip': ip_address,
        'username': username,
        'password': password,
    }

    # Connexion au commutateur
    connection = ConnectHandler(**device)

    # Vérification de la configuration actuelle pour l'accès Web
    output = connection.send_command("show running-config | include ip http")

    # Vérification si les lignes "ip http server" ou "ip http secure-server" sont présentes
    if "ip http server" in output or "ip http secure-server" in output:
        print(f"L'accès Web est activé sur {ip_address}. Désactivation en cours...")

        # Entrer en mode de configuration
        connection.config_mode()

        # Désactivation de l'accès Web
        connection.send_command("no ip http server")
        connection.send_command("no ip http secure-server")

        # Sauvegarde de la configuration
        connection.send_command("write memory")

        print(f"L'accès Web a été désactivé avec succès sur {ip_address}.")

        # Mettre à jour la colonne D avec "OK"
        sheet.cell(row=row_index, column=4, value="OK")

    else:
        print(f"L'accès Web n'est pas activé sur {ip_address}.")

        # Mettre à jour la colonne D avec "N/A"
        sheet.cell(row=row_index, column=4, value="N/A")

    # Déconnexion du commutateur
    connection.disconnect()

# Enregistrez les modifications dans le fichier Excel
workbook.save("inventaire_switch_0day_Copie.xlsx")

# Fermer le fichier Excel
workbook.close()
